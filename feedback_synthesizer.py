"""
Customer Feedback Synthesizer — AI Agent
Portfolio Project #1

Usage:
    python feedback_synthesizer.py reviews.csv

Input CSV columns (required): id, source, rating, text, date
Output: <name>_report.xlsx (3 sheets: Dashboard, Backlog, Raw Data)

Requires:
    pip install openai openpyxl
    export OPENAI_API_KEY=sk-...
"""

import sys
import os
import json
import csv
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ─────────────────────────────────────────────────────────────────────

MODEL = "gpt-4o"

COLORS = {
    "header_bg":  "1E1B4B",
    "accent":     "4F46E5",
    "accent_lt":  "EEF2FF",
    "critical":   "FEE2E2",
    "critical_t": "DC2626",
    "high":       "FEF3C7",
    "high_t":     "D97706",
    "medium":     "DBEAFE",
    "medium_t":   "2563EB",
    "low":        "DCFCE7",
    "low_t":      "16A34A",
    "row_alt":    "F8F7FF",
    "white":      "FFFFFF",
    "text_dark":  "1E293B",
    "subtext":    "64748B",
}

# ── HELPERS ────────────────────────────────────────────────────────────────────

def hex_fill(h):
    return PatternFill("solid", start_color=h, end_color=h)

def make_border():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def severity_colors(severity):
    return {
        "Critical": (COLORS["critical"], COLORS["critical_t"]),
        "High":     (COLORS["high"],     COLORS["high_t"]),
        "Medium":   (COLORS["medium"],   COLORS["medium_t"]),
        "Low":      (COLORS["low"],      COLORS["low_t"]),
    }.get(severity, (COLORS["white"], COLORS["text_dark"]))

def effort_colors(effort):
    return {
        "High":   (COLORS["critical"], COLORS["critical_t"]),
        "Medium": (COLORS["high"],     COLORS["high_t"]),
        "Low":    (COLORS["low"],      COLORS["low_t"]),
    }.get(effort, (COLORS["white"], COLORS["text_dark"]))

# ── STEP 1: READ CSV ───────────────────────────────────────────────────────────

def load_reviews(csv_path):
    reviews = []
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            reviews.append({
                "id":     int(row.get("id", 0)),
                "source": row.get("source", "Unknown"),
                "rating": int(row.get("rating", 3)),
                "text":   row.get("text", "").strip(),
                "date":   row.get("date", ""),
            })
    return reviews

# ── STEP 2: OPENAI CLUSTERING ──────────────────────────────────────────────────

SYSTEM_PROMPT = """You are a senior product manager specializing in user research and feedback analysis.
Your job is to analyze customer reviews and extract structured product opportunities.

Return ONLY valid JSON — no markdown, no explanation, just the JSON object."""

USER_PROMPT = """Analyze these {n} customer reviews for a SaaS product.

REVIEWS:
{reviews_json}

Return a JSON object with this exact structure:
{{
  "clusters": [
    {{
      "id": "C1",
      "theme": "Short theme name (3-5 words)",
      "description": "1-2 sentence description of the pattern across reviews",
      "review_ids": [list of review IDs belonging to this cluster],
      "frequency": number of reviews in this cluster,
      "avg_rating": average rating of reviews in this cluster (1 decimal),
      "severity": "Critical" | "High" | "Medium" | "Low",
      "severity_score": integer 1-10,
      "opportunity_score": integer 1-100,
      "effort": "High" | "Medium" | "Low",
      "recommended_action": "Specific actionable recommendation for the product team (2-3 sentences)",
      "user_story": "As a [user type], I want [goal] so that [benefit]",
      "kpi": "Metric to track success of this improvement",
      "sample_quote": "The single most representative verbatim quote from this cluster"
    }}
  ],
  "summary": {{
    "total_reviews": {n},
    "critical_count": number of Critical severity clusters,
    "avg_rating": overall average rating (1 decimal),
    "top_insight": "The single most important finding in 1-2 sentences"
  }}
}}

Rules:
- Create 5-8 clusters (don't over-fragment)
- Every review_id must appear in exactly one cluster
- Sort clusters by opportunity_score descending
- severity: Critical = causes churn/uninstall, High = blocks key action, Medium = creates friction, Low = minor polish
- opportunity_score = frequency + severity + strategic impact combined"""


def analyze_reviews(reviews, api_key):
    client = OpenAI(api_key=api_key)

    reviews_json = json.dumps([
        {"id": r["id"], "source": r["source"], "rating": r["rating"], "text": r["text"]}
        for r in reviews
    ], indent=2)

    prompt = USER_PROMPT.format(n=len(reviews), reviews_json=reviews_json)

    print(f"🤖 Analyzing {len(reviews)} reviews with {MODEL}...")

    response = client.chat.completions.create(
        model=MODEL,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content": prompt},
        ],
        temperature=0.2,
        response_format={"type": "json_object"},
    )

    result = json.loads(response.choices[0].message.content)
    print(f"✅ Identified {len(result['clusters'])} clusters")
    return result

# ── STEP 3: BUILD EXCEL ────────────────────────────────────────────────────────

def build_dashboard(wb, clusters, summary):
    ws = wb.create_sheet("📊 Dashboard")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "Customer Feedback Synthesizer — Opportunity Dashboard"
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = hex_fill(COLORS["header_bg"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Analysis of {summary['total_reviews']} reviews  ·  {len(clusters)} clusters  ·  Avg rating: {summary['avg_rating']}/5  ·  {summary['top_insight']}"
    ws["A2"].font = Font(name="Arial", size=9, color="A5B4FC")
    ws["A2"].fill = hex_fill(COLORS["header_bg"])
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    ws.row_dimensions[3].height = 10
    ws.row_dimensions[4].height = 50
    ws.row_dimensions[5].height = 20

    kpis = [
        ("A4:B4", str(summary["total_reviews"]),   "Total Reviews Analyzed"),
        ("C4:D4", str(len(clusters)),               "Opportunity Clusters"),
        ("E4:F4", str(summary["critical_count"]),   "Critical / High Issues"),
        ("G4:H4", f"{summary['avg_rating']} / 5.0", "Avg User Rating"),
    ]
    for rng, val, label in kpis:
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value = val
        c.font = Font(name="Arial", bold=True, size=22, color=COLORS["accent"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = hex_fill(COLORS["accent_lt"])
        lc_addr = rng.split(":")[0].replace("4", "5")
        rc_addr = rng.split(":")[1].replace("4", "5")
        ws.merge_cells(f"{lc_addr}:{rc_addr}")
        lc = ws[lc_addr]
        lc.value = label
        lc.font = Font(name="Arial", size=9, color=COLORS["subtext"])
        lc.alignment = Alignment(horizontal="center")

    ws.row_dimensions[6].height = 10
    headers = ["#", "Opportunity Theme", "Reviews", "Avg Rating", "Severity", "Effort", "Score", "Recommended Next Step"]
    widths  = [4,   35,                  10,        11,           11,         9,        10,      55]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(c)].width = w
        cell = ws.cell(row=7, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill = hex_fill(COLORS["accent"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = make_border()
    ws.row_dimensions[7].height = 28

    for i, cl in enumerate(clusters, 1):
        row = 7 + i
        ws.row_dimensions[row].height = 55
        bg = COLORS["row_alt"] if i % 2 == 0 else COLORS["white"]
        sev_bg, sev_fg = severity_colors(cl["severity"])
        eff_bg, eff_fg = effort_colors(cl["effort"])
        row_data = [i, cl["theme"], cl["frequency"], cl["avg_rating"],
                    cl["severity"], cl["effort"], cl["opportunity_score"], cl["recommended_action"]]
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=c_idx, value=val)
            cell.border = make_border()
            cell.alignment = Alignment(vertical="center", wrap_text=True,
                                       horizontal="center" if c_idx in (1,3,4,7) else "left")
            if c_idx == 5:
                cell.fill = hex_fill(sev_bg)
                cell.font = Font(name="Arial", bold=True, size=9, color=sev_fg)
            elif c_idx == 6:
                cell.fill = hex_fill(eff_bg)
                cell.font = Font(name="Arial", bold=True, size=9, color=eff_fg)
            elif c_idx == 7:
                cell.fill = hex_fill(bg)
                cell.font = Font(name="Arial", bold=True, size=11, color=COLORS["accent"])
            else:
                cell.fill = hex_fill(bg)
                cell.font = Font(name="Arial", size=9, color=COLORS["text_dark"])

    ws.freeze_panes = "A8"


def build_backlog(wb, clusters):
    ws = wb.create_sheet("🎯 Opportunity Backlog")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"] = "Structured Opportunity Backlog — Ready for Sprint Planning"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = hex_fill(COLORS["header_bg"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    for col, w in zip("ABCDEF", [8, 28, 55, 48, 36, 32]):
        ws.column_dimensions[col].width = w

    headers = ["Priority", "Theme", "User Story", "Recommended Action", "Success KPI", "Sample User Quote"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill = hex_fill(COLORS["accent"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = make_border()
    ws.row_dimensions[3].height = 28

    for i, cl in enumerate(clusters, 1):
        row = 3 + i
        ws.row_dimensions[row].height = 85
        bg = COLORS["row_alt"] if i % 2 == 0 else COLORS["white"]
        sev_bg, sev_fg = severity_colors(cl["severity"])
        row_data = [
            f"P{i}  |  {cl['opportunity_score']}",
            cl["theme"],
            cl["user_story"],
            cl["recommended_action"],
            cl["kpi"],
            f'"{cl["sample_quote"]}"',
        ]
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=c_idx, value=val)
            cell.border = make_border()
            cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
            if c_idx == 1:
                cell.fill = hex_fill(sev_bg)
                cell.font = Font(name="Arial", bold=True, size=9, color=sev_fg)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif c_idx == 2:
                cell.fill = hex_fill(bg)
                cell.font = Font(name="Arial", bold=True, size=10, color=COLORS["text_dark"])
            elif c_idx == 6:
                cell.fill = hex_fill(bg)
                cell.font = Font(name="Arial", italic=True, size=9, color=COLORS["subtext"])
            else:
                cell.fill = hex_fill(bg)
                cell.font = Font(name="Arial", size=9, color=COLORS["text_dark"])

    ws.freeze_panes = "A4"


def build_raw_data(wb, reviews, clusters):
    ws = wb.create_sheet("📋 Raw Data")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"] = "Raw Feedback — All Reviews with Cluster Assignment"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = hex_fill(COLORS["header_bg"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    for col, w in zip("ABCDEF", [6, 12, 10, 12, 22, 80]):
        ws.column_dimensions[col].width = w

    headers = ["ID", "Source", "Rating", "Date", "Cluster", "Review Text"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill = hex_fill(COLORS["accent"])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = make_border()
    ws.row_dimensions[3].height = 28

    review_cluster = {}
    for cl in clusters:
        for rid in cl["review_ids"]:
            review_cluster[rid] = cl["theme"]

    for i, review in enumerate(reviews):
        row = 4 + i
        ws.row_dimensions[row].height = 30
        bg = COLORS["row_alt"] if i % 2 == 0 else COLORS["white"]
        row_data = [review["id"], review["source"], review["rating"],
                    review["date"], review_cluster.get(review["id"], "Unclassified"), review["text"]]
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=c_idx, value=val)
            cell.border = make_border()
            cell.fill = hex_fill(bg)
            cell.alignment = Alignment(vertical="center", wrap_text=(c_idx == 6),
                                       horizontal="center" if c_idx in (1,2,3,4) else "left")
            cell.font = Font(name="Arial", size=9, color=COLORS["text_dark"])
            if c_idx == 3:
                rating_color = {1:"FEE2E2", 2:"FEF3C7", 3:"FEF9C3", 4:"DCFCE7", 5:"BBF7D0"}.get(val, bg)
                cell.fill = hex_fill(rating_color)
                cell.font = Font(name="Arial", bold=True, size=9, color=COLORS["text_dark"])

    ws.freeze_panes = "A4"


def build_excel(clusters, summary, reviews, output_path):
    wb = Workbook()
    wb.remove(wb.active)
    build_dashboard(wb, clusters, summary)
    build_backlog(wb, clusters)
    build_raw_data(wb, reviews, clusters)
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)
    print(f"📊 Report saved: {output_path}")

# ── MAIN ───────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python feedback_synthesizer.py reviews.csv")
        print("       OPENAI_API_KEY must be set as environment variable")
        sys.exit(1)

    csv_path = sys.argv[1]
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        print("❌ OPENAI_API_KEY not set.")
        print("   Run: export OPENAI_API_KEY=sk-...")
        sys.exit(1)

    print(f"📂 Loading: {csv_path}")
    reviews = load_reviews(csv_path)
    print(f"   {len(reviews)} reviews loaded")

    result   = analyze_reviews(reviews, api_key)
    clusters = result["clusters"]
    summary  = result["summary"]

    csv_name    = os.path.splitext(os.path.basename(csv_path))[0]
    output_path = os.path.join(os.path.dirname(os.path.abspath(csv_path)), f"{csv_name}_report.xlsx")
    build_excel(clusters, summary, reviews, output_path)
    print(f"\n✅ Done! Open: {output_path}")

if __name__ == "__main__":
    main()
